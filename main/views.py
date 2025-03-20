from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .forms import ShipmentForm, LoginForm, RegisterForm
from .models import Shipment
import openpyxl
import datetime
from django.db import IntegrityError
from django.contrib.auth.decorators import login_required
from django.db import transaction

@login_required(login_url='login')
def clear_database(request):
    if request.method == 'POST':
        with transaction.atomic():
            # Assuming 'Shipment' is the model you want to clear
            Shipment.objects.all().delete()
            messages.success(request, "All shipments have been successfully deleted.")
        return redirect('shipment_list')  # Redirect to a safe page after clearing the database
    else:
        messages.error(request, "Invalid request method.")
        return redirect('shipment_list')


def clear_messages(request):
    """Helper function to clear all messages from the request."""
    storage = messages.get_messages(request)
    for message in storage:
        pass
    storage.used = True

def index(request):
    """Redirect root URL to login page."""
    return redirect('login')

@login_required(login_url='login')
def home(request):
    """Home page view that requires login."""
    clear_messages(request)
    return render(request, 'main/home.html')

@login_required(login_url='login')
def add_shipment(request):
    # Get all unique clients from the database
    existing_clients = Shipment.objects.values_list('Claiming_Client', flat=True).distinct()

    if request.method == 'POST':
        form = ShipmentForm(request.POST)
        claim_no = request.POST.get('Claim_No')

        if Shipment.objects.filter(Claim_No=claim_no).exists():
            existing_shipment = Shipment.objects.get(Claim_No=claim_no)
            messages.warning(request, f'Duplicate claim number {claim_no}, consider editing the existing entry.')
            return render(request, 'main/add_shipment.html', {
                'form': form,
                'clients': existing_clients,
                'duplicate_claim_no': claim_no,
                'edit_shipment_id': existing_shipment.id
            })

        if form.is_valid():
            form.save()
            messages.success(request, 'Shipment added successfully.')
            return redirect('shipment_list')
        else:
            messages.error(request, 'Please correct the errors below.')

    else:
        form = ShipmentForm()

    return render(request, 'main/add_shipment.html', {
        'form': form,
        'clients': existing_clients
    })
@login_required(login_url='login')
def shipment_list(request):
    """List all shipments with the option to apply filters."""
    clear_messages(request)
    shipments = Shipment.objects.all()
    branches = Shipment.objects.values_list('Branch', flat=True).distinct().order_by('Branch')
    shipments = apply_filters(request, shipments)
    return render(request, 'main/shipment_list.html', {
        'shipments': shipments,
        'branches': branches,
    })

@login_required(login_url='login')
def edit_shipment(request, pk):
    """Edit an existing shipment."""
    clear_messages(request)
    shipment = get_object_or_404(Shipment, pk=pk)
    
    # Check if this is a cancel/keep original request
    if request.method == 'POST' and 'keep_original' in request.POST:
        messages.info(request, 'No changes were made to the shipment.')
        return redirect('shipment_list')
    
    if request.method == 'POST':
        form = ShipmentForm(request.POST, instance=shipment)
        if form.is_valid():
            form.save()
            messages.success(request, 'Shipment updated successfully.')
            return redirect('shipment_list')
    else:
        form = ShipmentForm(instance=shipment)
    return render(request, 'main/edit_shipment.html', {'form': form, 'shipment': shipment})

@login_required(login_url='login')
def export_shipments_excel(request):
    """Export shipment data to an Excel file."""
    clear_messages(request)
    shipments = Shipment.objects.all()
    response = create_excel_response(shipments)
    return response

@login_required(login_url='login')
def import_shipments(request):
    """Import shipment data from an Excel file, with detailed feedback on the process."""
    clear_messages(request)
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'Invalid file format. Please upload an Excel file.')
            return render(request, 'main/import_shipments.html')

        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            worksheet = wb.active
            skipped_entries, created_entries, error_entries = process_excel_data(worksheet)
            if created_entries == 0 and not skipped_entries and not error_entries:
                messages.info(request, 'No new entries were created. Check if the data is already up to date.')
            else:
                messages.success(request, f'Successfully created {created_entries} entries. Skipped {len(skipped_entries)} duplicate entries.')
                if error_entries:
                    messages.error(request, f'Errors occurred in {len(error_entries)} entries. {", ".join(error_entries)}')
        except Exception as e:
            messages.error(request, f'An error occurred while processing the file: {str(e)}')

        return render(request, 'main/import_shipments.html')
    return render(request, 'main/import_shipments.html')

@login_required(login_url='login')
def get_client_names(request):
    """Fetch client names dynamically based on user's query for autocomplete functionalities."""
    clear_messages(request)
    query = request.GET.get('q', '')
    clients = Shipment.objects.filter(Claiming_Client__icontains=query).distinct()
    return JsonResponse(list(clients), safe=False)

def user_login(request):
    if request.method == 'POST':
        form = LoginForm(request, data=request.POST)
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(username=username, password=password)
        if user:
            if user.is_active:
                login(request, user)
                request.session.set_expiry(0)  # Expire session on browser close
                return redirect('home')
        else:
            messages.error(request, "Invalid username or password. Please try again.")
    else:
        form = LoginForm()

    response = render(request, 'main/login.html', {'form': form})
    response['Cache-Control'] = 'no-store'
    return response
def register(request):
    """Handle new user registration and automatic login upon successful registration."""
    clear_messages(request)
    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            request.session.set_expiry(0)
            return redirect('home')
        else:
            messages.error(request, "Error in form submission.")
    else:
        form = RegisterForm()
    return render(request, 'main/register.html', {'form': form})

def user_logout(request):
    """Handle user logout and redirect to login page."""
    clear_messages(request)
    logout(request)  # Clears user session
    return redirect('login')

@login_required(login_url='login')
def delete_shipment(request, pk):
    """Delete a shipment entry."""
    clear_messages(request)
    shipment = get_object_or_404(Shipment, pk=pk)
    if request.method == 'POST':
        shipment.delete()
        messages.success(request, 'Shipment deleted successfully.')
    return redirect('shipment_list')

# Helper functions below
def apply_filters(request, shipments):
    """Apply filters to the shipments queryset based on request parameters."""
    
    # Filter by claim number
    claim_no = request.GET.get('claim_no')
    if claim_no:
        shipments = shipments.filter(Claim_No__icontains=claim_no)
    
    # Filter by client name
    client = request.GET.get('client')
    if client:
        shipments = shipments.filter(Claiming_Client__icontains=client)
    
    # Filter by branch
    branch = request.GET.get('branch')
    if branch:
        shipments = shipments.filter(Branch=branch)
    
    # Filter by Intend Date range
    intend_date_from = request.GET.get('intend_date_from')
    if intend_date_from:
        shipments = shipments.filter(Intend_Claim_Date__gte=intend_date_from)
    
    intend_date_to = request.GET.get('intend_date_to')
    if intend_date_to:
        shipments = shipments.filter(Intend_Claim_Date__lte=intend_date_to)
    
    # Filter by Formal Date range
    formal_date_from = request.GET.get('formal_date_from')
    if formal_date_from:
        shipments = shipments.filter(Formal_Claim_Date_Received__gte=formal_date_from)
    
    formal_date_to = request.GET.get('formal_date_to')
    if formal_date_to:
        shipments = shipments.filter(Formal_Claim_Date_Received__lte=formal_date_to)
    
    return shipments

def create_excel_response(shipments):
    # Create an Excel file response with shipment data
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Shipments'
    # Define headers and append them
    headers = ['Claim No', 'Claiming Client', 'Branch', 'Formal Claim', 'Intend Date', 
               'Formal Date', 'Claimed Amount', 'Amount Paid by Carrier', 
               'Amount Paid by AWA', 'Amount Paid by Insurance', 'Closed Date']
    worksheet.append(headers)
    # Append data rows
    for shipment in shipments:
        row = [
            shipment.Claim_No,
            shipment.Claiming_Client,
            shipment.Branch,
            shipment.Formal_Claim_Received,
            shipment.Intend_Claim_Date.strftime("%Y-%m-%d") if shipment.Intend_Claim_Date else '',
            shipment.Formal_Claim_Date_Received.strftime("%Y-%m-%d") if shipment.Formal_Claim_Date_Received else '',
            shipment.Claimed_Amount,
            shipment.Amount_Paid_By_Carrier,
            shipment.Amount_Paid_By_Awa,
            shipment.Amount_Paid_By_Insurance,
            shipment.Closed_Date.strftime("%Y-%m-%d") if shipment.Closed_Date else ''
        ]
        worksheet.append(row)
    # Create a response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="shipments.xlsx"'
    workbook.save(response)
    return response

def process_excel_data(worksheet):
    # Process Excel data and save valid entries to the database
    skipped_entries = []
    created_entries = 0
    error_entries = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:  # Skip empty rows or rows without Claim No
            continue
        claim_no = str(row[0])
        if Shipment.objects.filter(Claim_No=claim_no).exists():
            skipped_entries.append(claim_no)
            continue
        try:
            # Handle date conversions
            intend_date = None
            if row[4]:
                if isinstance(row[4], datetime.date):
                    intend_date = row[4]
                elif isinstance(row[4], str):
                    try:
                        intend_date = datetime.datetime.strptime(row[4], "%Y-%m-%d").date()
                    except ValueError:
                        # Try other date formats if the first one fails
                        try:
                            intend_date = datetime.datetime.strptime(row[4], "%d/%m/%Y").date()
                        except ValueError:
                            intend_date = None
            
            formal_date = None
            if row[5]:
                if isinstance(row[5], datetime.date):
                    formal_date = row[5]
                elif isinstance(row[5], str):
                    try:
                        formal_date = datetime.datetime.strptime(row[5], "%Y-%m-%d").date()
                    except ValueError:
                        # Try other date formats if the first one fails
                        try:
                            formal_date = datetime.datetime.strptime(row[5], "%d/%m/%Y").date()
                        except ValueError:
                            formal_date = None
            
            closed_date = None
            if row[10]:
                if isinstance(row[10], datetime.date):
                    closed_date = row[10]
                elif isinstance(row[10], str):
                    try:
                        closed_date = datetime.datetime.strptime(row[10], "%Y-%m-%d").date()
                    except ValueError:
                        # Try other date formats if the first one fails
                        try:
                            closed_date = datetime.datetime.strptime(row[10], "%d/%m/%Y").date()
                        except ValueError:
                            closed_date = None
            
            # Handle numeric conversions safely
            claimed_amount = 0
            if row[6]:
                if isinstance(row[6], (int, float)):
                    claimed_amount = float(row[6])
                elif isinstance(row[6], str) and row[6].strip():
                    # Remove any non-numeric characters except decimal point
                    clean_str = ''.join(c for c in row[6] if c.isdigit() or c == '.')
                    if clean_str:
                        claimed_amount = float(clean_str)
            
            carrier_amount = 0
            if row[7]:
                if isinstance(row[7], (int, float)):
                    carrier_amount = float(row[7])
                elif isinstance(row[7], str) and row[7].strip():
                    clean_str = ''.join(c for c in row[7] if c.isdigit() or c == '.')
                    if clean_str:
                        carrier_amount = float(clean_str)
            
            awa_amount = 0
            if row[8]:
                if isinstance(row[8], (int, float)):
                    awa_amount = float(row[8])
                elif isinstance(row[8], str) and row[8].strip():
                    clean_str = ''.join(c for c in row[8] if c.isdigit() or c == '.')
                    if clean_str:
                        awa_amount = float(clean_str)
            
            insurance_amount = 0
            if row[9]:
                if isinstance(row[9], (int, float)):
                    insurance_amount = float(row[9])
                elif isinstance(row[9], str) and row[9].strip():
                    clean_str = ''.join(c for c in row[9] if c.isdigit() or c == '.')
                    if clean_str:
                        insurance_amount = float(clean_str)
            
            # Convert formal claim received to YES/NO format
            formal_claim = "NO"
            if row[3]:
                if isinstance(row[3], bool):
                    formal_claim = "YES" if row[3] else "NO"
                elif isinstance(row[3], str):
                    if row[3].upper() in ["YES", "Y", "TRUE", "1"]:
                        formal_claim = "YES"
            
            # Get branch value and validate it's in BRANCH_CHOICES
            branch = row[2] or ""
            # Ensure branch code is valid (you may want to add more validation)
            if branch and branch not in [choice[0] for choice in Shipment.BRANCH_CHOICES]:
                branch = ""  # Set to empty if invalid
            
            shipment = Shipment(
                Claim_No=claim_no,
                Claiming_Client=row[1] or "",
                Branch=branch,
                Formal_Claim_Received=formal_claim,
                Intend_Claim_Date=intend_date,
                Formal_Claim_Date_Received=formal_date,
                Claimed_Amount=claimed_amount,
                Amount_Paid_By_Carrier=carrier_amount,
                Amount_Paid_By_Awa=awa_amount,
                Amount_Paid_By_Insurance=insurance_amount,
                Closed_Date=closed_date
            )
            shipment.save()
            created_entries += 1
        except Exception as e:
            error_entries.append(f'Row with Claim No {claim_no}: {str(e)}')

    return skipped_entries, created_entries, error_entries

@login_required(login_url='login')
def client_autocomplete(request):
    term = request.GET.get('term', '')
    clients = Shipment.objects.filter(
        Claiming_Client__icontains=term
    ).values_list('Claiming_Client', flat=True).distinct()
    
    suggestions = [{'id': client, 'text': client} for client in clients if client]

    return JsonResponse(suggestions, safe=False)    